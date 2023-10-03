import telebot
from telebot import types
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pyqtgraph as pg
import sys
import os
import subprocess
import pathlib
import dateparser
import sqlite3
import datetime

#токен телеграм бота
bot = telebot.TeleBot('6541675758:AAEQXEe2bSJKCtNeeu-FhntkrdSa-CpU_WU')

#номер пациента
number = 0

#класс в котором хранятся все люди проходящие анкетирование
class People(object):
    
    #инициализация всех атрибутов класса
    def __init__(self, number):
        """Constructor"""
        self.user_id = 0
        
        self.number = number
        
        self.first_name = "Отсутствует"
        self.last_name = "Отсутствует"
        self.patronymic = "Отсутствует"
        
        self.gender = "Отсутствует"

        self.birth_day = 0
        self.birth_month = 0
        self.birth_year = 0

        self.health = "Отсутствует"
        self.infectious_diseases = "Отсутствует"
        self.heart_disease = "Отсутствует"
        self.allergic_reactions  = "Отсутствует"
        self.diseases_nervous_system  = "Отсутствует"
        self.diabetes_mellitus  = "Отсутствует"
        self.contact_infectious_diseases  = "Отсутствует"
        self.sexual_relations  = "Отсутствует"
        self.threat_of_epidemics  = "Отсутствует"
        self.narcotic_drugs  = "Отсутствует"
        self.vaccination  = "Отсутствует"
        self.medicines  = "Отсутствует"
        self.alcohol  = "Отсутствует"
        self.dispensary_registration  = "Отсутствует"
        self.piercing  = "Отсутствует"

#функция для отображения парметров пациетнта
def print_people(num):
    
    print("Ваша анкета:\nВаш номер: " +  str(Patient[num].number) +
            "\nВаш id: " +  str(Patient[num].user_id) +
            "\nВаша фамилия: " +  str(Patient[num].last_name) +
            "\nВаше имя: " +  str(Patient[num].first_name) +
            "\nВаше отчество: " +  str(Patient[num].patronymic) +
            "\nВаше пол: " +  str(Patient[num].gender) +
            "\nВаша дата рождения: " + str(Patient[num].birth_day) + '.'  +  str(Patient[num].birth_month) + "." +str(Patient[num].birth_year) +
            "\nХорошее ли у Вас самочувствие: " +  str(Patient[num].health) +
            "\nБыли ли у Вас когда-либо инфекционные заболевания: " +  str(Patient[num].infectious_diseases) +
            "\nБыли ли у Вас когда-либо болезни сердца: " +  str(Patient[num].heart_disease) +
            "\nБыли ли у Вас когда-либо тяжелые аллергические реакции, бронхиальная астма: " +  str(Patient[num].allergic_reactions) +
            "\nБыли ли у Вас когда-либо судороги и заболевания нервной системы: " +  str(Patient[num].diseases_nervous_system) +    
            "\nБыли ли у Вас когда-либо сахарный диабет, онкологические заболевания: " +  str(Patient[num].diabetes_mellitus) +     
            "\nНаходились ли Вы в контакте с больными инфекционными заболеваниями: " +  str(Patient[num].contact_infectious_diseases) +    
            "\nБыли ли у Вас сексуальные связи с лицами, инфицированными ВИЧ-инфекцией, больными вирусными гепатитами B и C, сифилисом: " +  str(Patient[num].sexual_relations) +            
            "\nПребывали ли Вы на территориях, на которых существует угроза возникновения массовых инфекционных заболеваний или эпидемий: " +  str(Patient[num].threat_of_epidemics) +            
            "\nУпотребляли ли Вы когда-либо наркотические средства, психотропные вещества: " +  str(Patient[num].narcotic_drugs) +            
            "\nПроводилась ли Вам за последний год вакцинация (прививки) и хирургические вмешательства: " +  str(Patient[num].vaccination) +            
            "\nПринимаете ли Вы в настоящее время или принимали в течение последних 30 календарных дней какие-либо лекарства, включая жаропонижающие: " +  str(Patient[num].medicines) +            
            "\nПринимали ли Вы за последние 48 часов алкоголь: " +  str(Patient[num].alcohol) +            
            "\nСостоите ли Вы на диспансерном учете или наблюдаетесь сейчас у врача: " +  str(Patient[num].dispensary_registration) +
            "\nПроводили ли Вам иглоукалывание, пирсинг, татуировку за последние 120 календарных дней: " +  str(Patient[num].piercing))

#функция для отображения парметров пациетнта
def string_people(num):
    
    s = ("Ваша анкета:\nВаш номер: " +  str(Patient[num].number) +
            "\nВаш id: " +  str(Patient[num].user_id) +
            "\nВаша фамилия: " +  str(Patient[num].last_name) +
            "\nВаше имя: " +  str(Patient[num].first_name) +
            "\nВаше отчество: " +  str(Patient[num].patronymic) +
            "\nВаше пол: " +  str(Patient[num].gender) +
            "\nВаша дата рождения: " + str(Patient[num].birth_day) + '.'  +  str(Patient[num].birth_month) + "." +str(Patient[num].birth_year) +
            "\nХорошее ли у Вас самочувствие: " +  str(Patient[num].health) +
            "\nБыли ли у Вас когда-либо инфекционные заболевания: " +  str(Patient[num].infectious_diseases) +
            "\nБыли ли у Вас когда-либо болезни сердца: " +  str(Patient[num].heart_disease) +
            "\nБыли ли у Вас когда-либо тяжелые аллергические реакции, бронхиальная астма: " +  str(Patient[num].allergic_reactions) +
            "\nБыли ли у Вас когда-либо судороги и заболевания нервной системы: " +  str(Patient[num].diseases_nervous_system) +    
            "\nБыли ли у Вас когда-либо сахарный диабет, онкологические заболевания: " +  str(Patient[num].diabetes_mellitus) +     
            "\nНаходились ли Вы в контакте с больными инфекционными заболеваниями: " +  str(Patient[num].contact_infectious_diseases) +    
            "\nБыли ли у Вас сексуальные связи с лицами, инфицированными ВИЧ-инфекцией, больными вирусными гепатитами B и C, сифилисом: " +  str(Patient[num].sexual_relations) +            
            "\nПребывали ли Вы на территориях, на которых существует угроза возникновения массовых инфекционных заболеваний или эпидемий: " +  str(Patient[num].threat_of_epidemics) +            
            "\nУпотребляли ли Вы когда-либо наркотические средства, психотропные вещества: " +  str(Patient[num].narcotic_drugs) +            
            "\nПроводилась ли Вам за последний год вакцинация (прививки) и хирургические вмешательства: " +  str(Patient[num].vaccination) +            
            "\nПринимаете ли Вы в настоящее время или принимали в течение последних 30 календарных дней какие-либо лекарства, включая жаропонижающие: " +  str(Patient[num].medicines) +            
            "\nПринимали ли Вы за последние 48 часов алкоголь: " +  str(Patient[num].alcohol) +            
            "\nСостоите ли Вы на диспансерном учете или наблюдаетесь сейчас у врача: " +  str(Patient[num].dispensary_registration) +
            "\nПроводили ли Вам иглоукалывание, пирсинг, татуировку за последние 120 календарных дней: " +  str(Patient[num].piercing))
    return s



@bot.message_handler(content_types=['text'])

#стартовое сообщение с которого начинается работа бота
def start(message):
    global number
    print(message.from_user.id)
    if message.text == '/start':
        number = number + 1
        Patient[number].user_id = message.from_user.id
        bot.send_message(message.from_user.id, "Введите Ваше ФИО:")
        bot.register_next_step_handler(message, get_ser_name)
        
    else:
        bot.send_message(message.from_user.id, 'Напиши /start')

def get_ser_name(message):
    try:
        st = message.text
        st = st.split()
    except:
        st = 'Иванов Иван Иванович'

    num = get_num_of_patient(message)
    
    try:
        Patient[num].last_name = st[0].capitalize()
    except:
        Patient[num].last_name = ' '
        
    try:
        Patient[num].first_name = st[1].capitalize()
    except:
        Patient[num].first_name = ' '

    try:
        Patient[num].patronymic = st[2].capitalize()
    except:
        Patient[num].patronymic = ' '
        

    question = ('Ваша фамилия: '+str(Patient[num].last_name)
                +'\nВаше имя: '+str(Patient[num].first_name)+'\nВаше отчетсво: '+str(Patient[num].patronymic)+'\nВсе верно?')
    keyboard = new_keyboard('Да', 'yes', 'Нет', 'no')
    bot.send_message(message.from_user.id, text=question, reply_markup=keyboard)


def get_data(message):
    try:
        s = dateparser.parse(message.text, settings={'DATE_ORDER': 'DMY'})
        day = s.day
        month = s.month
        year = s.year
    except:
        day = 1
        month = 1
        year = 1900

    num = get_num_of_patient(message)
    
    Patient[num].birth_day = day
    Patient[num].birth_month = month
    Patient[num].birth_year = year

    if(len(str(Patient[num].birth_month))==1):
        m1 = '0'+str(Patient[num].birth_month)
    else:
        m1 = str(Patient[num].birth_month)

    if(len(str(Patient[num].birth_day))==1):
        d1 = '0'+str(Patient[num].birth_day)
    else:
        d1 = str(Patient[num].birth_day)
    
    question = ('Дата вашего рождения: '+ d1 + '.' + m1 +'.'+str(Patient[num].birth_year) + '\nВерно?')
    keyboard = new_keyboard('Да', 'yes4', 'Нет', 'no4')
    bot.send_message(message.from_user.id, text=question, reply_markup=keyboard)   

@bot.callback_query_handler(func=lambda call: True)

#функция для работы с кнопками
def callback_worker(call):
    num = get_num_of_patient(call)
    
    #обработка ФИО
    btn_handler_for_two_btn(call, "yes", 'Введите дату вашего рождения', get_data, "no",'Повторно введите Ваше ФИО',get_ser_name)

    #обработка дня рождения
    if call.data == "yes4":
        keyboard = new_keyboard('Мужской', 'man', 'Женский', 'women','Предыдущий вопрос','previous1')
        bot.send_message(call.message.chat.id, text = "Укажите ваш пол", reply_markup=keyboard)

    elif call.data == "no4":
        bot.send_message(call.message.chat.id, 'Повторно введите дату вашего рождения')
        bot.register_next_step_handler(call.message, get_data)
    
    #опрос пола
    elif call.data == 'man':
        num = get_num_of_patient(call)
        Patient[num].gender = 'Мужской'
        print('Мужской')
        
        keyboard = new_keyboard('Да', 'yes5', 'Нет', 'no5','Предыдущий вопрос','previous5')
        bot.send_message(call.message.chat.id, text = "Хорошее ли у Вас сейчас самочувствие?", reply_markup=keyboard)
        
    elif call.data == 'women':
        num = get_num_of_patient(call)
        Patient[num].gender = 'Женский'
        print('Женский')
        print_people(num)

        keyboard = new_keyboard('Да', 'yes5', 'Нет', 'no5','Предыдущий вопрос','previous5')
        bot.send_message(call.message.chat.id, text = "Хорошее ли у Вас сейчас самочувствие?", reply_markup=keyboard)
        
    elif call.data == 'previous1':
        bot.send_message(call.message.chat.id, 'Повторно введите день вашего рождения числом')
        bot.register_next_step_handler(call.message, get_day)

    #опрос самочувствия
    elif call.data == 'yes5':
        Patient[num].health = 'Да'

        keyboard = new_keyboard('Да', 'yes6', 'Нет', 'no6','Предыдущий вопрос','previous6')
        bot.send_message(call.message.chat.id,
                         text = "Были ли у Вас когда-либо инфекционные заболевания, вирусные гепатиты B и C, сифилис, туберкулез, малярия)?",
                         reply_markup=keyboard);
        
    elif call.data == 'no5':
        Patient[num].health = 'Нет'

        keyboard = new_keyboard('Да', 'yes6', 'Нет', 'no6','Предыдущий вопрос','previous6')
        bot.send_message(call.message.chat.id,
                         text = "Были ли у Вас когда-либо инфекционные заболевания, вирусные гепатиты B и C, сифилис, туберкулез, малярия)?",
                         reply_markup=keyboard)
        
    elif call.data == 'previous5':
        keyboard = new_keyboard('Мужской', 'man', 'Женский', 'women','Предыдущий вопрос','previous1')
        bot.send_message(call.message.chat.id, text = "Повторно укажите пол", reply_markup=keyboard)

    #опрос инфекционных заболевапний
    Patient[num].infectious_diseases = btn_handler_for_three(call, Patient[num].infectious_diseases, 'yes6', 'no6', 'previous6',
                                        'yes7', 'no7', 'previous7', "Были ли у Вас когда-либо болезни сердца, высокое или низкое артериальное давление?",
                                        'yes5', 'no5', 'previous5', "Хорошее ли у Вас сейчас самочувствие?")

    #опрос болезний сердца
    Patient[num].heart_disease = btn_handler_for_three(call, Patient[num].heart_disease, 'yes7', 'no7', 'previous7',
                                        'yes8', 'no8', 'previous8', "Были ли у Вас когда-либо тяжелые аллергические реакции, бронхиальная астма?",
                                        'yes6', 'no6', 'previous6', "Были ли у Вас когда-либо инфекционные заболевания, вирусные гепатиты B и C, сифилис, туберкулез, малярия)?")

    #опрос на аллергические реакции
    Patient[num].allergic_reactions = btn_handler_for_three(call, Patient[num].allergic_reactions, 'yes8', 'no8', 'previous8',
                                        'yes9', 'no9', 'previous9', "Были ли у Вас когда-либо судороги и заболевания нервной системы?",
                                        'yes7', 'no7', 'previous7', "Были ли у Вас когда-либо болезни сердца, высокое или низкое артериальное давление?")

    #опрос на заболевания нервной системы
    Patient[num].diseases_nervous_system = btn_handler_for_three(call, Patient[num].diseases_nervous_system, 'yes9', 'no9', 'previous9',
                                        'yes10', 'no10', 'previous10', "Были ли у Вас когда-либо сахарный диабет, онкологические заболевания?",
                                        'yes8', 'no8', 'previous8', "Были ли у Вас когда-либо тяжелые аллергические реакции, бронхиальная астма?")

    #опрос на сахарный диабет
    Patient[num].diabetes_mellitus = btn_handler_for_three(call, Patient[num].diabetes_mellitus, 'yes10', 'no10', 'previous10',
                                        'yes11', 'no11', 'previous11', "Находились ли Вы в контакте с больными инфекционными заболеваниями?",
                                        'yes9', 'no9', 'previous9', "Были ли у Вас когда-либо судороги и заболевания нервной системы?")

    #опрос на контакт с больными инфекционными заболеваниями
    Patient[num].contact_infectious_diseases = btn_handler_for_three(call, Patient[num].contact_infectious_diseases, 'yes11', 'no11', 'previous11',
                                        'yes12', 'no12', 'previous12', "Были ли у Вас сексуальные связи с лицами, инфицированными вирусом иммунодефицита человека (ВИЧ-инфекцией), больными вирусными гепатитами B и C, сифилисом?",
                                        'yes10', 'no10', 'previous10', "Были ли у Вас когда-либо сахарный диабет, онкологические заболевания?")

    #опрос на сексуальные связи
    Patient[num].sexual_relations = btn_handler_for_three(call, Patient[num].sexual_relations, 'yes12', 'no12', 'previous12',
                                        'yes13', 'no13', 'previous13', "Пребывали ли Вы на территориях, на которых существует угроза возникновения и (или) распространения массовых инфекционных заболеваний или эпидемий?",
                                        'yes11', 'no11', 'previous11',"Находились ли Вы в контакте с больными инфекционными заболеваниями?")

    #опрос на угрозу возникновения эпидемий
    Patient[num].threat_of_epidemics = btn_handler_for_three(call, Patient[num].threat_of_epidemics, 'yes13', 'no13', 'previous13',
                                        'yes14', 'no14', 'previous14', "Употребляли ли Вы когда-либо наркотические средства, психотропные вещества?",
                                        'yes12', 'no12', 'previous12',"Были ли у Вас сексуальные связи с лицами, инфицированными вирусом иммунодефицита человека (ВИЧ-инфекцией), больными вирусными гепатитами B и C, сифилисом?")

    #опрос на наркотические средства
    Patient[num].narcotic_drugs = btn_handler_for_three(call, Patient[num].narcotic_drugs, 'yes14', 'no14', 'previous14',
                                        'yes15', 'no15', 'previous15', "Проводилась ли Вам за последний год вакцинация (прививки) и хирургические вмешательства?",
                                        'yes13', 'no13', 'previous13',"Пребывали ли Вы на территориях, на которых существует угроза возникновения и (или) распространения массовых инфекционных заболеваний или эпидемий?")

    #опрос на вакцинацию
    Patient[num].vaccination = btn_handler_for_three(call, Patient[num].vaccination, 'yes15', 'no15', 'previous15',
                                        'yes16', 'no16', 'previous16', "Принимаете ли Вы в настоящее время или принимали в течение последних 30 календарных дней какие-либо лекарства, включая жаропонижающие?",
                                        'yes14', 'no14', 'previous14',"Употребляли ли Вы когда-либо наркотические средства, психотропные вещества?")

    #опрос на лекарства
    Patient[num].medicines = btn_handler_for_three(call, Patient[num].medicines, 'yes16', 'no16', 'previous16',
                                        'yes17', 'no17', 'previous17', "Принимали ли Вы за последние 48 часов алкоголь?",
                                        'yes15', 'no15', 'previous15',"Проводилась ли Вам за последний год вакцинация (прививки) и хирургические вмешательства?")

    #опрос на алкоголь
    Patient[num].alcohol = btn_handler_for_three(call, Patient[num].alcohol, 'yes17', 'no17', 'previous17',
                                        'yes18', 'no18', 'previous18', "Состоите ли Вы на диспансерном учете или наблюдаетесь сейчас у врача?",
                                        'yes16', 'no16', 'previous16',"Принимаете ли Вы в настоящее время или принимали в течение последних 30 календарных дней какие-либо лекарства, включая жаропонижающие?")
  
    #опрос на диспансерный учет
    Patient[num].dispensary_registration = btn_handler_for_three(call, Patient[num].dispensary_registration, 'yes18', 'no18', 'previous18',
                                        'yes19', 'no19', 'previous19', "Проводили ли Вам иглоукалывание, пирсинг, татуировку за последние 120 календарных дней?",
                                        'yes17', 'no17', 'previous17',"Принимали ли Вы за последние 48 часов алкоголь?")
    
    #опрос пирсинг
    if call.data == 'yes19':
        num = get_num_of_patient(call)
        answer = 'Да'
        q = string_people(num)
        bot.send_message(call.message.chat.id, text = q)
        keyboard = new_keyboard('Да', 'yes20', 'Нет, начать сначала', 'no20','Предыдущий вопрос', 'previous20')
        bot.send_message(call.message.chat.id, text = "Сохранить анкету", reply_markup=keyboard)
        Patient[num].piercing = answer
        
    elif call.data == 'no19':
        num = get_num_of_patient(call)
        answer = 'Нет'
        q = string_people(num)
        bot.send_message(call.message.chat.id, text = q)
        keyboard = new_keyboard('Да', 'yes20', 'Нет, начать сначала', 'no20','Предыдущий вопрос', 'previous20')
        bot.send_message(call.message.chat.id, text = "Сохранить анкету", reply_markup=keyboard)
        
        Patient[num].piercing = answer
        

    elif call.data == 'previous19':
        keyboard = new_keyboard('Да', 'yes18', 'Нет', 'no18','Предыдущий вопрос','previous18')
        bot.send_message(call.message.chat.id, text = "Состоите ли Вы на диспансерном учете или наблюдаетесь сейчас у врача?", reply_markup=keyboard)



    #последний вопрос
    elif call.data == 'yes20':
        #сохранение в эксель
        save_in_exel(num)

        #сохранение в sqlite
        save_in_sqlite_bd(num)
        
        answer = ("Ваша анкета сохранена")
        bot.send_message(call.message.chat.id, text = answer)

    elif call.data == 'no20':
        bot.send_message(call.message.chat.id, 'Повторно введите Ваше ФИО')
        bot.register_next_step_handler(call.message, get_ser_name)

    elif call.data == 'previous20':
        keyboard = new_keyboard('Да', 'yes19', 'Нет', 'no19','Предыдущий вопрос','previous19')
        bot.send_message(call.message.chat.id, text = "Проводили ли Вам иглоукалывание, пирсинг, татуировку за последние 120 календарных дней?", reply_markup=keyboard)
   
    print_people(num)

#сохранение данных в эксель
def save_in_exel(num):
    wb = Workbook()
    wb = load_workbook('./anketa_donora.xlsx')
    sheet = wb.active

    
    sheet["G2"] = str(Patient[num].last_name)+" "+str(Patient[num].first_name)+" "+str(Patient[num].patronymic)

    if(len(str(Patient[num].birth_month))==1):
        m1 = '0'+str(Patient[num].birth_month)
    else:
        m1 = str(Patient[num].birth_month)

    if(len(str(Patient[num].birth_day))==1):
        d1 = '0'+str(Patient[num].birth_day)
    else:
        d1 = str(Patient[num].birth_day)
    
    sheet["G3"] = d1 + '.'  +  m1 + "." +str(Patient[num].birth_year)
    sheet["L6"] = str(Patient[num].health)
    sheet["L7"] = str(Patient[num].infectious_diseases)
    sheet["L8"] = str(Patient[num].heart_disease)
    sheet["L9"] = str(Patient[num].allergic_reactions)
    sheet["L10"] = str(Patient[num].diseases_nervous_system)
    sheet["L11"] = str(Patient[num].diabetes_mellitus)
    sheet["L12"] = str(Patient[num].contact_infectious_diseases)
    sheet["L13"] = str(Patient[num].sexual_relations)
    sheet["L14"] = str(Patient[num].threat_of_epidemics)
    sheet["L15"] = str(Patient[num].narcotic_drugs)
    sheet["L16"] = str(Patient[num].vaccination)
    sheet["L17"] = str(Patient[num].medicines)
    sheet["L18"] = str(Patient[num].alcohol)
    sheet["L19"] = str(Patient[num].dispensary_registration)
    sheet["L20"] = str(Patient[num].piercing)
    #имя файла эксель из ФИО и id
    file_name=str(Patient[num].last_name)+" "+str(Patient[num].first_name)+" "+str(Patient[num].patronymic)+" "+str(Patient[num].user_id)
    p=str(pathlib.Path.cwd())
    p=p.replace('\\', '/')
    wb.save(p+'/donors/'+file_name+'.xlsx')

def save_in_sqlite_bd(num):
    # Устанавливаем соединение с базой данных
    connection = sqlite3.connect('database.db')

    #создаем объект "курсор" для выполнения SQL-запросов и операций с базой данных
    cursor = connection.cursor()

    # Создаем таблицу Patient
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS Patient (
    Время_создания TEXT PRIMARY KEY,
    id TEXT NOT NULL,
    ФИО TEXT NOT NULL,
    Дата_рождения TEXT NOT NULL,
    Q1 TEXT NOT NULL,
    Q2 TEXT NOT NULL,
    Q3 TEXT NOT NULL,
    Q4 TEXT NOT NULL,
    Q5 TEXT NOT NULL,
    Q6 TEXT NOT NULL,
    Q7 TEXT NOT NULL,
    Q8 TEXT NOT NULL,
    Q9 TEXT NOT NULL,
    Q10 TEXT NOT NULL,
    Q11 TEXT NOT NULL,
    Q12 TEXT NOT NULL,
    Q13 TEXT NOT NULL,
    Q14 TEXT NOT NULL,
    Q15 TEXT NOT NULL
    )
    ''')


    name = str(Patient[num].last_name)+" "+str(Patient[num].first_name)+" "+str(Patient[num].patronymic)

    if(len(str(Patient[num].birth_month))==1):
        m1 = '0'+str(Patient[num].birth_month)
    else:
        m1 = str(Patient[num].birth_month)

    if(len(str(Patient[num].birth_day))==1):
        d1 = '0'+str(Patient[num].birth_day)
    else:
        d1 = str(Patient[num].birth_day)
    
    data = d1 + '.'  +  m1 + "." +str(Patient[num].birth_year)

    dt_now = datetime.datetime.now()
    # Добавляем нового пользователя
    cursor.execute('INSERT INTO Patient (Время_создания, id, ФИО, Дата_рождения, Q1,Q2,Q3,Q4,Q5,Q6,Q7,Q8,Q9,Q10,Q11,Q12,Q13,Q14,Q15) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                    (str(dt_now), str(Patient[num].user_id), name, data, str(Patient[num].health),str(Patient[num].infectious_diseases),
                    str(Patient[num].heart_disease),str(Patient[num].allergic_reactions),
                    str(Patient[num].diseases_nervous_system),str(Patient[num].diabetes_mellitus),
                    str(Patient[num].contact_infectious_diseases),str(Patient[num].sexual_relations),
                    str(Patient[num].threat_of_epidemics),str(Patient[num].narcotic_drugs),
                    str(Patient[num].vaccination),str(Patient[num].medicines),
                    str(Patient[num].alcohol),str(Patient[num].dispensary_registration),str(Patient[num].piercing)))

    


    # Сохраняем изменения и закрываем соединение
    connection.commit()
    connection.close()

#быстрое создание клавиатуры
def new_keyboard(*btns):
    keyboard = types.InlineKeyboardMarkup()
    for i in range(0,len(btns),2):
        key_yes = types.InlineKeyboardButton(text=btns[i], callback_data=btns[i+1])
        keyboard.add(key_yes)
    return keyboard

#быстрое создание диалогового акна с двумя кнопками с переходами
def btn_handler_for_two_btn(call, btn1, text1, func1, btn2,text2,func2):
    if call.data == btn1:
        bot.send_message(call.message.chat.id, text1)
        bot.register_next_step_handler(call.message, func1)
        
    elif call.data == btn2:
        bot.send_message(call.message.chat.id, text2)
        bot.register_next_step_handler(call.message, func2)

#быстрое создание диалогового акна с тремя кнопками с переходами
def btn_handler_for_three(call, var, btn1, btn2, btn3,
                                        new_btn1, new_btn2, new_btn3, new_text,
                                        old_btn1, old_btn2, old_btn3, old_text):
    
    if call.data == btn1:
        num = get_num_of_patient(call)
        answer = 'Да'
        keyboard = new_keyboard('Да', new_btn1, 'Нет', new_btn2,'Предыдущий вопрос', new_btn3)
        bot.send_message(call.message.chat.id, text = new_text, reply_markup=keyboard)
        return answer
        

    elif call.data == btn2:
        num = get_num_of_patient(call)
        answer = 'Нет'

        keyboard = new_keyboard('Да', new_btn1, 'Нет', new_btn2,'Предыдущий вопрос', new_btn3)
        bot.send_message(call.message.chat.id, text = new_text, reply_markup=keyboard)
        return answer
        

    elif call.data == btn3:
        keyboard = new_keyboard('Да', old_btn1, 'Нет', old_btn2,'Предыдущий вопрос',old_btn3)
        bot.send_message(call.message.chat.id, text = old_text, reply_markup=keyboard)
        return var

    return var

#получение номера поциента по его id
def get_num_of_patient(message):
    for i in range(len(Patient)):
        if(Patient[i].user_id == message.from_user.id):
            num = i
    return num


#создание списка из 100 экземпляров класса
Patient = []
for i in range(100):
    Patient.append(People(i))
    
#для экселя
openpyxl

#запуск бота
bot.polling(none_stop=True, interval=0)
